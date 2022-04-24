from tkinter import font
from openpyxl import Workbook
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment, PatternFill
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
import src.packages.othersrc.constants as cts
from src.packages.othersrc.genFunc import checkKeyInDict, createDir, handleError, handleInput, displayLine

class ExcelHandler():

    def __init__(self):
        self.wb=Workbook()
        self.styles={}
        self.__setSheetStyle(self.wb.active)
        self.__setDefaultStyles()
    
    def __setSheetStyle(self, ws):
        ws.sheet_properties.tabColor="FFFFFF"
        ws.sheet_view.showGridLines=cts.SHOW_GRIDLINES

    def __setDefaultStyles(self):
        head=NamedStyle(name="head")
        head.font=Font(name='Calibri',size=11,bold=True)
        head.border=Border(top=Side(style='thin',color='000000'), bottom=Side(style='thick',color='000000'))
        self.wb.add_named_style(head)
        self.styles["head"]=head

        total=NamedStyle(name="total")
        total.font=Font(name='Calibri',size=11,bold=True)
        total.border=Border(top=Side(style='thin',color='000000'), bottom=Side(style='double',color='000000'))
        self.wb.add_named_style(total)
        self.styles["total"]=total
            
        bold=NamedStyle(name="bold")
        bold.font=Font(name='Calibri',size=11,bold=True)
        self.wb.add_named_style(bold)
        self.styles["bold"]=bold
        
        bad=NamedStyle(name="bad")
        bad.font=Font(name='Calibri',size=11,bold=True,color='9C0006')
        bad.fill=PatternFill("solid", fgColor='FFC7CE')
        self.wb.add_named_style(bad)
        self.styles["bad"]=bad
        
        good=NamedStyle(name="good")
        good.font=Font(name='Calibri',size=11,bold=True,color='006100')
        good.fill=PatternFill("solid", fgColor='C6EFCE')
        self.wb.add_named_style(good)
        self.styles["good"]=good
        
    def writeToCell(self, ws, colName, rowNum, writeVal=""):
        ws[f"{colName}{rowNum}"]=writeVal
        
    def writeToRange(self, ws, fromCell, toCell, writeArr=[], rangeNamedStyle="", fontSize=0, numberFormat=""):
        cellRange=ws[fromCell:toCell]
        verticalSelection=True
        if(fromCell[0]!=toCell[0]):
            verticalSelection=False
            cellRange=cellRange[0]
            
        if(not len(writeArr)==len(cellRange)):
            return
        
        for i in range(len(cellRange)):
            if(verticalSelection):
                cell=cellRange[i][0]
            else:
                cell=cellRange[i]
            cell.value=writeArr[i]
            self.cellFormatNumber(ws, f"{cell.coordinate}", numberFormat)
            if(rangeNamedStyle):
                try:
                    cell.style=self.styles[rangeNamedStyle]
                    if(fontSize):
                        try:
                            cell.font=Font(size=fontSize)
                        except:
                            pass
                except:
                    pass
            
    def setCellStyle(self, ws, cell, style):
        ws[cell].style=self.styles[style]

    def createSheet(self, sheetName, position=None):
        if(not position):
            ws=self.wb.create_sheet(sheetName, position)
        else:    
            ws=self.wb.create_sheet(sheetName)

        self.__setSheetStyle(ws)

        return ws

    def saveWorkbook(self, wbName=""):
        if(not wbName):
            wbName="./" + cts.OUTPUT_FOLDER + "/" + cts.WORKBOOK_NAME
        elif(not wbName.__contains__(cts.OUTPUT_FOLDER)):
            wbName="./" + cts.OUTPUT_FOLDER + "/" + wbName
            
        if(checkKeyInDict(self.wb, "Sheet", False)):
            try:
                self.wb.remove("Sheet")
            except:
                pass
            
        createDir("./" + cts.OUTPUT_FOLDER)
        
        try:
            displayLine("Guardando Excel {}...".format(wbName))
            self.wb.save(wbName)
            displayLine("Guardando exitoso")
        except:
            handleError("Acceso al archivo {} denegado".format(wbName))
            userIn=handleInput("Nuevo nombre del Excel (enter para el mismo): ")
            if(userIn):
                if(not userIn.__contains__(".")):
                    userIn += ".xlsx"
                return self.saveWorkbook("./" + cts.OUTPUT_FOLDER + "/" + userIn)
            else:
                return self.saveWorkbook(wbName)
        handleInput(f"Excel guardado en {wbName}. Presionar una tecla para cerrar")

    def cellFormatNumber(cls, ws, cell, format="A"):
        match format.upper():
            case "P":
                ws["{}".format(cell)].number_format="0.{}%".format("0" * cts.ROUND_DECIMALS)
            case "A":
                ws["{}".format(cell)].number_format=cts.NUMBER_FORMAT_ACCOUNTING
        return ""

    def conditionalFormatting(self, ws, cell, cutoutNumber=0):
        ws.conditional_formatting.add(cell, CellIsRule(operator='lessThanOrEqual', formula=[f"{cutoutNumber}"], font=Font(name='Calibri',size=11,bold=True,color='9C0006'), fill=PatternFill("solid", start_color='FFC7CE', end_color='FFC7CE')))    
        ws.conditional_formatting.add(cell, CellIsRule(operator='greaterThan', formula=[f"{cutoutNumber}"], font=Font(name='Calibri',size=11,bold=True,color='006100'), fill=PatternFill("solid", start_color='C6EFCE', end_color='C6EFCE')))